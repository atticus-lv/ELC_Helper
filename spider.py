#-*- coding =utf-8 -*-

from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options

import time,os
import sqlite3
import openpyxl


def get_ele_info():
    student = UserInfo()
    username, password = student.get()
    driver = getdriver()

    if driver:
        web = check_electricity(username, password, driver)
        if web.login():
            time_str,year,month,day = get_cn_time()
            dictionary = web.get_info_list()
            if dictionary:
                print("数据获取完毕")
                return dictionary,time_str


def deal_info():
    dict, time_str = get_ele_info()
    msg = time_str +"\n>>宿舍电量小助手<<"

    for key,value in dict.items():
        str = key + "：" + value
        msg = msg +"\n" + str
    return msg



def get_cn_time():
    daylist = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    CNdaylist = ["日", "一", "二", "三", "四", "五", "六"]

    year = time.strftime("%Y", time.localtime())
    month = time.strftime("%m", time.localtime())
    day = time.strftime("%d", time.localtime())
    week = time.strftime("%A", time.localtime())
    now = time.strftime("%H:%M", time.localtime())

    i = daylist.index(week)
    str = f"今天是{year}年{month}月{day}日，星期{CNdaylist[i]}"

    return str,year,month,day


def getdriver():
    try:

        chrome_options = Options()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('log-level=3')
        print("驱动正常")
        print("静默运行模式...")
        driver = webdriver.Chrome('chromedriver.exe',options=chrome_options)

        return driver
    except:
        print("目录文件夹下缺少驱动程序，无法使用"
              "\n请检查浏览器驱动版本是否正确,本文件所附带驱动版本为 Chrome 84"
              "\n下载正确版本：http://chromedriver.storage.googleapis.com/index.html")
        return None


class check_electricity(object):
    def __init__(self,username,password,driver):
        self.username = username
        self.password = password
        self.driver = driver

    def login(self):
        print("开始寻找网页")
        self.driver.get(
            "http://kl.bnuz.edu.cn/")
        self.driver.implicitly_wait(10)
        print("到达指定页面，开始填写用户数据")
        self.driver.find_element_by_xpath('//*[@id="ctl00_cphContent_txtUsername"]').send_keys(
            self.username)
        self.driver.find_element_by_xpath('//*[@id="ctl00_cphContent_txtPassword"]').send_keys(
            self.password)
        self.driver.find_element_by_xpath('//*[@id="ctl00_cphContent_btnLogin"]').click()
        return True

    def get_info_list(self):
        print("正在获取数据")
        nameList = ["用户名","地址","剩余电量","购电单价"]
        infoList = []

        address1 = self.driver.find_element_by_xpath('//*[@id="ctl00_cphContent_lblUsername"]')
        address2 = self.driver.find_element_by_xpath('//*[@id="ctl00_cphContent_lblAddress"]')
        dianliang = self.driver.find_element_by_xpath('//*[@id="ctl00_cphContent_LabelSY"]')
        danjia = self.driver.find_element_by_xpath('//*[@id="ctl00_cphContent_lblprice"]')

        rawinfo = [address1,address2,dianliang,danjia]
        for info in rawinfo:
            infoList.append(info.text)

        dictionary = dict(zip(nameList, infoList))
        return dictionary


class UserInfo(object):

    @staticmethod
    def read():
        print("欢迎使用电费查询小助手")
        print("检测账号密码文件中...")
        f = open("账户密码.txt", "r")  # 打开文件 w为写入，没有则创建
        lines = f.readlines()  # readline 为一行 readlines为全部行，单独read为单个字符位
        username = lines[0]
        password = lines[1]
        f.close()  # 关闭文件'''
        return username, password

    @staticmethod
    def write():
        print("没有检测到账号密码文件")
        print("请依次输入账号密码")
        username = input("账号")
        password = input("密码")
        info = open("账户密码.txt", "w")
        info.write(username)
        info.write("\n")
        info.write(password)
        info.close()
        print('账号密码文件生成中...\n下次可直接登录')
        return username, password

    @classmethod
    def get(self):
        try:
            username, password = self.read()
            print("读取文件中...")
            return username, password
        except:
            username, password = self.write()
            return username, password
        finally:
            print('开始检测浏览器驱动')


def init_db(dbpath):  # 初始化数据库
    # 创建创建数据库
    sql = '''         
        create table ele_data
        (
        id integer primary key autoincrement,
        data text,
        value numeric
        )

    '''
    # 创建数据表
    conn = sqlite3.connect(dbpath)  #
    cursor = conn.cursor()
    cursor.execute(sql)
    conn.commit()  # 提交
    conn.close()  # 关闭


def creat_new_file(path):
    data = openpyxl.Workbook()  # 新建工作簿
    data.create_sheet('宿舍用电数据')  # 添加页
    table = data.get_sheet_by_name('宿舍用电数据') # 获得指定名称页
    table = data.active  # 获得当前活跃的工作页，默认为第一个工作页
     # 行，列，值 这里是从1开始计数的
    table.cell(1, 1).value = "时间"
    table.cell(1, 2).value="电量"
    table.cell(1, 3).value = "当日用电量"
    data.save(path)


def read_latest(path):
    data = openpyxl.load_workbook(path)
    table = data.get_sheet_by_name('宿舍用电数据')
    table = data.active
    nrows = table.max_row  # 获得行数
    ncols = table.max_column  # 获得列数
    old_value = table.cell(nrows-1,ncols-1).value
    return  nrows,ncols,data,old_value


def write_latest(nrows,ncols,data,old_value,time_today,latest_value):
    data = openpyxl.load_workbook(path)
    table = data.get_sheet_by_name('宿舍用电数据')
    table = data.active
    #时间
    table.cell(nrows,1 ).value = time_today
    # 新数据
    table.cell(nrows,2).value = latest_value
    #差量
    try:
        print(latest_value, old_value)
        value = float(latest_value) - float(old_value)

        table.cell(nrows-1,ncols ).value = value
    except Exception as e:
        print(e)

    data.save(path)

if __name__ == '__main__':

    dictionary, time_str =get_ele_info()
    latest_value = dictionary["剩余电量"]
    str, year, month, day = get_cn_time()
    time_today = f"{year}/{month}/{day}"

    path = "Data/用电数据.xlsx"

    # creat_new_file(path)
    nrows,ncols,data,old_value = read_latest(path)
    # print(nrows,ncols,data,old_value )
    write_latest(nrows,ncols,data,old_value,time_today,latest_value)


    # print(deal_info())
    time.sleep(5)
